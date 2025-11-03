# routers/dashboard.py
from datetime import datetime, timedelta
import random
from fastapi import APIRouter, Request, Query 
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse
from fastapi.templating import Jinja2Templates

from services.database import get_db_connection
from services.state import connected_clients
from utils.utility_functions import template_response 
import pandas as pd 
from itertools import groupby 
from openpyxl import load_workbook  
from openpyxl.styles import Alignment 
import os 

router = APIRouter() 
templates = Jinja2Templates(directory="templates")

# @router.get("/", response_class=HTMLResponse)
# async def home(request: Request, message: str = None):
#     return templates.TemplateResponse(
#         "login.html",{"request": request} 
#     ) 
#     # return RedirectResponse(url="/Dashboard")  


@router.get("/Dashboard", response_class=HTMLResponse)
async def dashboard(request: Request, message: str = None):
    message = request.query_params.get("message")
    conn = get_db_connection()
    installed_meters = conn.execute("SELECT * FROM installed_meters").fetchall() 
    registered_meter_count = conn.execute("SELECT COUNT(*) FROM registered_meters").fetchone()[0]  
    total_installations_count = conn.execute("SELECT COUNT(*) FROM installed_meters").fetchone()[0] 
    type_stats = conn.execute(
        """
        SELECT type, COUNT(*) as count 
        FROM installed_meters
        WHERE type IN ('DDSY283SR', 'DTSD546', 'DTSD545S')
        GROUP BY type
        """
    ).fetchall()
    type_counts = {row["type"]: row["count"] for row in type_stats}

    # Ensure missing types are set to 0
    for t in ["DDSY283SR", "DTSD546", "DTSD545S"]:
        type_counts.setdefault(t, 0)
    print(type_counts)

    conn.close()


    return template_response(request, "dashboard.html", 
        {
            "request": request,
            "message": message,
            "registered_meters": installed_meters, 
            "total_installations": total_installations_count, 
            "total_online_meter": len(connected_clients),
            "type_counts": type_counts,
            "registered_meter_count":registered_meter_count,
        }) 
    # return templates.TemplateResponse(
    #     "dashboard.html", 
    #     {
    #         "request": request,
    #         "message": message,
    #         "registered_meters": installed_meters, 
    #         "total_installations": total_installations_count, 
    #         "total_online_meter": len(connected_clients),
    #         "online_rate":  len(connected_clients)/total_installations_count * 100, 
    #         "type_counts": type_counts,
    #         "registered_meter_count":registered_meter_count,
    #     }
    # )


@router.get("/active-power-last6h", response_class=HTMLResponse) 
async def getActivePowerdata(request: Request,
                    line: str = "", 
                    start: str = "", 
                    end: str = ""  
                    ):
    column_name = "total_active_power"  
    table_name = "regular_task_readings"  
    
    query = f"""
        SELECT meter_number, timestamp, {column_name}  
        FROM {table_name} 
        WHERE 1 = 1
    """
    params = [] 
    meters = get_meter_by_line(line) 
    meter_data = {}
    all_timestamps_sets = []  

    for meter in meters:  
        if meter:  
            query += " AND meter_number LIKE ?"
            params.append(f"%{meter}%")  
        if start and end:
            query += " AND timestamp BETWEEN ? AND ?"
            params.extend([start, end]) 
        
        conn = get_db_connection()
        readings = conn.execute(query, params).fetchall() 
        conn.close()
        readings_dict = {round_to_minute(datetime.fromisoformat(r['timestamp'])): r[f'{column_name}'] for r in readings} 
 
        meter_data[meter] = readings_dict 
        all_timestamps_sets.append(set(readings_dict.keys())) 

    # Step 3: Find common timestamps
    common_timestamps = set.intersection(*all_timestamps_sets) 


    total_load = []
    for ts in sorted(common_timestamps):
        total_value = sum(meter_data[m][ts] for m in meters) 
        total_load.append({
            "timestamp": ts.isoformat(),
            "value": total_value
        })
    return JSONResponse(total_load) 




 
@router.get("/hourly-consumption-last24h", response_class=HTMLResponse) 
async def getConsumptionData24(request: Request,
                    line: str = "", 
                    start: str = "", 
                    end: str = ""  
                    ):
    column_name = "import_total_active_energy"  
    table_name = "energy_profile_readings_calculated"  
    
    query = f"""
        SELECT meter_number, timestamp, {column_name}  
        FROM {table_name} 
        WHERE 1 = 1
    """
    params = [] 
    meters = get_meter_by_line(line) 
    meter_data = {}
    all_timestamps_sets = [] 

    for meter in meters:  
        if meter:  
            query += " AND meter_number LIKE ?"
            params.append(f"%{meter}%")  
        if start and end:
            query += " AND timestamp BETWEEN ? AND ?"
            params.extend([start, end]) 
        
        conn = get_db_connection()
        readings = conn.execute(query, params).fetchall() 
        conn.close()
        readings_dict = {round_to_hour(datetime.fromisoformat(r['timestamp'])): r[f'{column_name}'] for r in readings} 
 
        meter_data[meter] = readings_dict 
        all_timestamps_sets.append(set(readings_dict.keys())) 

    # Step 3: Find common timestamps
    common_timestamps = set.intersection(*all_timestamps_sets) 
    sorted_hours = sorted(common_timestamps) 
    meter_hourly_values = {} 
    for meter in meters:
        readings = meter_data[meter]
        # Filter readings only for common timestamps
        filtered = {ts: readings[ts] for ts in sorted_hours if ts in readings}
        # Sorted list of (ts, value)
        meter_hourly_values[meter] = sorted(filtered.items())  



    hourly_consumptions = []  
    for i in range(1, len(sorted_hours)):
        ts_prev = sorted_hours[i - 1]
        ts_curr = sorted_hours[i]

        total_diff = 0
        # Sum differences across all meters
        for meter in meters:
            prev_value = dict(meter_hourly_values[meter]).get(ts_prev)
            curr_value = dict(meter_hourly_values[meter]).get(ts_curr)

            # Ensure both readings exist
            if prev_value is not None and curr_value is not None:
                diff = curr_value - prev_value
                # You might want to ignore negative diffs or treat as zero if your meter resets
                total_diff += max(diff, 0)

        hourly_consumptions.append({
            "timestamp": ts_curr.isoformat(),
            "value": total_diff
        })
    return JSONResponse(hourly_consumptions) 


@router.get("/daily-consumption-last30d", response_class=HTMLResponse)
async def getConsumptionData30(
    request: Request,
    line: str = "", 
    start: str = "", 
    end: str = ""
):
    column_name = "import_total_active_energy"  
    table_name = "energy_profile_readings_calculated"   
    print(start) 
    print(end) 
    # If no date range is provided, default to last 30 days
    if not start or not end:
        end_dt = datetime.now()
        start_dt = end_dt - timedelta(days=30)
        start = start_dt.strftime("%Y-%m-%d 00:00:00")
        end = end_dt.strftime("%Y-%m-%d 23:59:59")
    
    query = f"""
        SELECT meter_number, timestamp, {column_name}  
        FROM {table_name} 
        WHERE 1 = 1
    """
    params = [] 
    meters = get_meter_by_line(line) 
    meter_data = {}
    all_timestamps_sets = [] 

    for meter in meters:  
        if meter:  
            query += " AND meter_number LIKE ?"
            params.append(f"%{meter}%")  
        if start and end:
            query += " AND timestamp BETWEEN ? AND ?"
            params.extend([start, end]) 
        
        conn = get_db_connection()
        readings = conn.execute(query, params).fetchall() 
        conn.close()

        # 00:00:00 tsagiin zaaltuudiig shuuj avah
        readings_dict = {datetime.fromisoformat(r['timestamp']).replace(hour=0, minute=0, second=0, microsecond=0): 
                         r[f'{column_name}'] for r in readings} 
 
        meter_data[meter] = readings_dict 
        all_timestamps_sets.append(set(readings_dict.keys())) 

    # Find common days
    common_days = set.intersection(*all_timestamps_sets) 
    sorted_days = sorted(common_days) 

    meter_daily_values = {} 
    for meter in meters:
        readings = meter_data[meter]
        filtered = {ts: readings[ts] for ts in sorted_days if ts in readings}
        meter_daily_values[meter] = sorted(filtered.items())  

    # Calculate daily consumption
    daily_consumptions = []  
    for i in range(1, len(sorted_days)):
        day_prev = sorted_days[i - 1]
        day_curr = sorted_days[i]

        total_diff = 0
        for meter in meters:
            prev_value = dict(meter_daily_values[meter]).get(day_prev)
            curr_value = dict(meter_daily_values[meter]).get(day_curr)
            if prev_value is not None and curr_value is not None:
                diff = curr_value - prev_value
                total_diff += max(diff, 0)

        daily_consumptions.append({
            "timestamp": day_curr.strftime("%Y-%m-%d"),
            "value": total_diff
        })

    return JSONResponse(daily_consumptions)




def get_meter_by_line(
    line: str = "" 
    ):
    query = "SELECT meter_number FROM installed_meters WHERE 1=1"  
    params = []
    
    if line is not None:   
        query+= " AND line LIKE ?"   
        params.append(f"%{line}%")  
    conn = get_db_connection()
    meters = [row[0] for row in conn.execute(query, params).fetchall()] 
    print(meters) 
    conn.close() 
    return meters


def round_to_minute(ts):  # oirhon minut luu shiljuuleh 
    """Round timestamp to the nearest minute."""
    return ts.replace(second=0, microsecond=0)  
def round_to_hour(dt):    # oirhon tsag ruu horvuuleh  
    return dt.replace(minute=0, second=0, microsecond=0)  

def get_lines():
    conn = get_db_connection()
    meters = conn.execute("SELECT * FROM installed_meters").fetchall()   
    conn.close()
    line_names = [meter['line'] for meter in meters] 
    return line_names
def get_meter_numbers(): 
    conn = get_db_connection()
    meters = conn.execute("SELECT * FROM installed_meters").fetchall()   
    conn.close()
    meter_numbers = [meter['meter_number'] for meter in meters]  
    return meter_numbers 



@router.get("/export-Excel-24H-consumpsion")
def export_excel_24h_consumption(selected_date: str = Query(...)):
    print("📅 Selected date:", selected_date)

    # 1️⃣ Parse the selected date
    date_obj = datetime.strptime(selected_date, "%Y-%m-%d")

    # 2️⃣ Generate 24 hourly intervals for that date
    hours_list = [f"{hour:02d}:00" for hour in range(24)]
    hours_list1 = [
        (date_obj + timedelta(hours=h)).strftime("%Y-%m-%dT%H:00:00")
        for h in range(24)
    ]
    print("🕒 Hours list:", hours_list1)

    # 3️⃣ Your existing code with hours_list1 instead of last 24h
    lines = get_lines()
    meter_numbers = get_meter_numbers()
    obis_codes = ["1.8.0", "3.8.0"]
    table_name = "energy_profile_readings_calculated"
    data = []

    conn = get_db_connection()
    cursor = conn.cursor()

    for line, meter in zip(lines, meter_numbers):
        for obis in obis_codes:
            if obis == "1.8.0":
                column_name = "import_total_active_energy"
            else:
                column_name = "import_total_reactive_energy"

            query = f"""
                SELECT meter_number, timestamp, {column_name}
                FROM {table_name}
                WHERE meter_number = ? AND timestamp >= ? AND timestamp < ?
            """

            for hour in hours_list1:
                start_time = datetime.fromisoformat(hour)
                end_time = start_time + timedelta(hours=1)
                start_time_str = start_time.strftime("%Y-%m-%dT%H:%M:%S")
                end_time_str = end_time.strftime("%Y-%m-%dT%H:%M:%S")

                cursor.execute(query, (meter, start_time_str, end_time_str))
                row = cursor.fetchone()
                value = row[2] if row else None

                data.append([line, meter, obis, start_time.strftime("%H:00"), value])

    conn.close()

    # 4️⃣ Convert to DataFrame and pivot
    df = pd.DataFrame(data, columns=["Line", "Meter Number", "OBIS", "Hour", "Value"])
    pivot_df = df.pivot_table(
        index=["Line", "Meter Number", "OBIS"],
        columns="Hour",
        values="Value",
        aggfunc="first"
    ).reset_index()

    # Reorder columns to keep hours in order
    existing_hours = [h for h in hours_list if h in pivot_df.columns]
    pivot_df = pivot_df[["Line", "Meter Number", "OBIS"] + existing_hours]

    # 5️⃣ Export to Excel
    file_path = f"temp_export_{selected_date}.xlsx"
    pivot_df.to_excel(file_path, index=False)
    wb = load_workbook(file_path)
    ws = wb.active

    def merge_same_cells(col_index):
        values = [ws.cell(row=i, column=col_index).value for i in range(2, ws.max_row+1)]
        for key, group in groupby(enumerate(values, start=2), lambda x: x[1]):
            group = list(group)
            if key is not None and len(group) > 1:
                start_row = group[0][0]
                end_row = group[-1][0]
                ws.merge_cells(start_row=start_row, start_column=col_index,
                               end_row=end_row, end_column=col_index)
                ws.cell(row=start_row, column=col_index).alignment = Alignment(horizontal="center", vertical="center")
            elif key is not None:
                ws.cell(row=group[0][0], column=col_index).alignment = Alignment(horizontal="center", vertical="center")

    merge_same_cells(1)
    merge_same_cells(2)
    wb.save(file_path)

    return FileResponse(
        file_path,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        filename=os.path.basename(file_path)
    )
